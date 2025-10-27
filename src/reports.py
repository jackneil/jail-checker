"""
Report generators for custody check results.

This module provides functions to generate custody reports in JSON and Excel formats.
"""

import json
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
import logging
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from models import CustodyReport, CustodyResult

# Configure logging
logger = logging.getLogger(__name__)


def generate_json_report(report: CustodyReport, output_path: str | Path) -> Path:
    """
    Generate a JSON report of custody check results.

    >>> from models import Defendant, CustodyResult, CustodyReport
    >>> import tempfile
    >>> import os
    >>> defendants = [Defendant(last_name="Smith", first_name="John")]
    >>> results = [CustodyResult(defendant_name="Smith, John", in_custody=True)]
    >>> report = CustodyReport(defendants_checked=defendants, custody_results=results)
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     output = generate_json_report(report, os.path.join(tmpdir, "report.json"))
    ...     output.exists()
    True

    Args:
        report: CustodyReport object
        output_path: Path where JSON file should be saved

    Returns:
        Path object pointing to the generated JSON file
    """
    output_path = Path(output_path)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Convert report to dictionary
    report_dict = {
        "search_date": report.search_date.isoformat(),
        "source_file": report.source_file,
        "summary": {
            "total_defendants": report.total_defendants,
            "in_custody": report.in_custody_count,
            "not_in_custody": report.not_in_custody_count,
            "errors": report.error_count
        },
        "defendants": [
            {
                "full_name": d.full_name,
                "first_name": d.first_name,
                "middle_name": d.middle_name,
                "last_name": d.last_name,
                "matter_number": d.matter_number,
                "case_number": d.case_number,
                "charges": d.charges,
                "incident_date": d.incident_date,
                "case_status": d.case_status
            }
            for d in report.defendants_checked
        ],
        "custody_results": [
            {
                "defendant_name": r.defendant_name,
                "matter_number": r.matter_number,
                "case_number": r.case_number,
                "in_custody": r.in_custody,
                "booking_number": r.booking_number,
                "booking_date": r.booking_date,
                "custody_location": r.custody_location,
                "charges_at_booking": r.charges_at_booking,
                "bond_amount": r.bond_amount,
                "mugshot_url": r.mugshot_url,
                "query_timestamp": r.query_timestamp.isoformat(),
                "error_message": r.error_message,
                "status_summary": r.status_summary
            }
            for r in report.custody_results
        ],
        "in_custody_list": [
            {
                "defendant_name": r.defendant_name,
                "matter_number": r.matter_number,
                "case_number": r.case_number,
                "booking_date": r.booking_date,
                "charges": r.charges_at_booking,
                "bond_amount": r.bond_amount
            }
            for r in report.get_in_custody_list()
        ]
    }

    # Write JSON file
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report_dict, f, indent=2, ensure_ascii=False)

    return output_path


def generate_excel_report(report: CustodyReport, output_path: str | Path) -> Path:
    """
    Generate an Excel report of custody check results.

    >>> from models import Defendant, CustodyResult, CustodyReport
    >>> import tempfile
    >>> import os
    >>> defendants = [Defendant(last_name="Smith", first_name="John")]
    >>> results = [CustodyResult(defendant_name="Smith, John", in_custody=True)]
    >>> report = CustodyReport(defendants_checked=defendants, custody_results=results)
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     output = generate_excel_report(report, os.path.join(tmpdir, "report.xlsx"))
    ...     output.exists()
    True

    Args:
        report: CustodyReport object
        output_path: Path where Excel file should be saved

    Returns:
        Path object pointing to the generated Excel file
    """
    output_path = Path(output_path)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create Summary sheet
    _create_summary_sheet(wb, report)

    # Create All Results sheet
    _create_all_results_sheet(wb, report)

    # Create In Custody sheet
    _create_in_custody_sheet(wb, report)

    # Save workbook
    wb.save(output_path)

    return output_path


def _download_mugshot(url: str) -> Optional[bytes]:
    """
    Download a mugshot image from URL and return as bytes.

    Args:
        url: URL to mugshot image

    Returns:
        Image bytes or None if download fails
    """
    if not url:
        return None

    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return response.content
        else:
            logger.warning(f"Failed to download mugshot: {url} (status {response.status_code})")
            return None
    except Exception as e:
        logger.warning(f"Error downloading mugshot from {url}: {str(e)}")
        return None


def _create_mugshot_image(image_bytes: bytes, width: int = 100, height: int = 125) -> Image:
    """
    Create a fresh openpyxl Image object from image bytes.

    Must be called separately for each cell to avoid corruption.

    Args:
        image_bytes: Raw image data
        width: Desired width in pixels (default 100)
        height: Desired height in pixels (default 125)

    Returns:
        Fresh openpyxl Image object
    """
    image_data = BytesIO(image_bytes)
    image_data.seek(0)
    img = Image(image_data)
    img.width = width
    img.height = height
    return img


def _download_mugshots_parallel(results: List[CustodyResult]) -> Dict[str, bytes]:
    """
    Download all mugshots for in-custody defendants in parallel.

    Args:
        results: List of CustodyResult objects

    Returns:
        Dictionary mapping mugshot URL to image bytes
    """
    # Collect unique mugshot URLs for in-custody defendants
    mugshot_urls = set()
    for result in results:
        if result.in_custody and result.mugshot_url:
            mugshot_urls.add(result.mugshot_url)

    if not mugshot_urls:
        logger.info("No mugshots to download")
        return {}

    logger.info(f"Downloading {len(mugshot_urls)} mugshots in parallel...")

    images = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        # Submit all download tasks
        future_to_url = {
            executor.submit(_download_mugshot, url): url
            for url in mugshot_urls
        }

        # Collect results as they complete
        completed = 0
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                image_bytes = future.result()
                if image_bytes:
                    images[url] = image_bytes
                    completed += 1
                    if completed % 5 == 0:
                        logger.info(f"Downloaded {completed}/{len(mugshot_urls)} mugshots")
            except Exception as e:
                logger.error(f"Error downloading {url}: {str(e)}")

    logger.info(f"Successfully downloaded {len(images)}/{len(mugshot_urls)} mugshots")
    return images


def _create_summary_sheet(wb: Workbook, report: CustodyReport):
    """
    Create the Summary sheet in the Excel workbook.

    Args:
        wb: Workbook object
        report: CustodyReport object
    """
    ws = wb.create_sheet("Summary")

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)

    # Add title
    ws['A1'] = "Custody Check Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')

    # Add report info
    row = 3
    info_items = [
        ("Report Date:", report.search_date.strftime('%Y-%m-%d %H:%M:%S')),
        ("Source File:", report.source_file or "N/A"),
        ("", ""),
        ("Total Defendants Checked:", str(report.total_defendants)),
        ("In Custody:", str(report.in_custody_count)),
        ("Not in Custody:", str(report.not_in_custody_count)),
        ("Errors:", str(report.error_count)),
    ]

    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label:  # Bold the labels
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40


def _create_all_results_sheet(wb: Workbook, report: CustodyReport):
    """
    Create the All Results sheet in the Excel workbook with mugshots.

    Args:
        wb: Workbook object
        report: CustodyReport object
    """
    ws = wb.create_sheet("All Results")

    # Download all mugshots in parallel first
    mugshot_images = _download_mugshots_parallel(report.custody_results)

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers (Mugshot column added as first column)
    headers = [
        "Mugshot",
        "Matter Number",
        "Case Number",
        "Defendant Name",
        "Custody Status",
        "Booking Date",
        "Booking Number",
        "Charges",
        "Bond Amount",
        "Error Message"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    in_custody_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
    not_in_custody_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green
    error_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Light yellow

    # Sort results: IN CUSTODY first, then NOT IN CUSTODY
    sorted_results = sorted(report.custody_results, key=lambda r: (not r.in_custody, r.defendant_name))

    for row_num, result in enumerate(sorted_results, 2):
        # Set row height for mugshots
        ws.row_dimensions[row_num].height = 100

        # Insert mugshot if available (column A)
        if result.in_custody and result.mugshot_url and result.mugshot_url in mugshot_images:
            # Create a fresh Image object for this row (can't reuse same Image)
            img = _create_mugshot_image(mugshot_images[result.mugshot_url])
            img.anchor = f'A{row_num}'
            ws.add_image(img)

        # Shift all data columns right by 1 (B, C, D, etc.)
        ws.cell(row=row_num, column=2, value=result.matter_number)
        ws.cell(row=row_num, column=3, value=result.case_number)
        ws.cell(row=row_num, column=4, value=result.defendant_name)
        ws.cell(row=row_num, column=5, value=result.status_summary)
        ws.cell(row=row_num, column=6, value=result.booking_date)
        ws.cell(row=row_num, column=7, value=result.booking_number)
        ws.cell(row=row_num, column=8, value=result.charges_at_booking)
        ws.cell(row=row_num, column=9, value=result.bond_amount)
        ws.cell(row=row_num, column=10, value=result.error_message)

        # Apply row coloring based on status
        if result.error_message:
            fill = error_fill
        elif result.in_custody:
            fill = in_custody_fill
        else:
            fill = not_in_custody_fill

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 16  # Mugshot (100px wide images)
    ws.column_dimensions['B'].width = 15  # Matter Number
    ws.column_dimensions['C'].width = 20  # Case Number
    ws.column_dimensions['D'].width = 30  # Defendant Name
    ws.column_dimensions['E'].width = 25  # Custody Status
    ws.column_dimensions['F'].width = 15  # Booking Date
    ws.column_dimensions['G'].width = 15  # Booking Number
    ws.column_dimensions['H'].width = 40  # Charges
    ws.column_dimensions['I'].width = 15  # Bond Amount
    ws.column_dimensions['J'].width = 30  # Error Message

    # Freeze header row
    ws.freeze_panes = 'A2'


def _create_in_custody_sheet(wb: Workbook, report: CustodyReport):
    """
    Create the In Custody sheet in the Excel workbook with mugshots.

    Args:
        wb: Workbook object
        report: CustodyReport object
    """
    ws = wb.create_sheet("In Custody Only")

    # Get in custody list first
    in_custody_list = report.get_in_custody_list()

    # Download all mugshots in parallel
    mugshot_images = _download_mugshots_parallel(in_custody_list)

    # Title row (merged across all columns)
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "DEFENDANTS IN CUSTODY"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Empty row
    ws.row_dimensions[2].height = 5

    # Header style
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Dark red
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers at row 3 (Mugshot added as first column)
    headers = [
        "Mugshot",
        "Matter Number",
        "Case Number",
        "Defendant Name",
        "Booking Date",
        "Booking Number",
        "Charges",
        "Bond Amount",
        "Original Charges",
        "Case Status"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows start at row 4 - only in custody
    highlight_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red

    for row_num, result in enumerate(in_custody_list, 4):
        # Set row height for mugshots
        ws.row_dimensions[row_num].height = 100

        # Insert mugshot if available (column A)
        if result.mugshot_url and result.mugshot_url in mugshot_images:
            # Create a fresh Image object for this row (can't reuse same Image)
            img = _create_mugshot_image(mugshot_images[result.mugshot_url])
            img.anchor = f'A{row_num}'
            ws.add_image(img)

        # Shift all data columns right by 1 (B, C, D, etc.)
        ws.cell(row=row_num, column=2, value=result.matter_number)
        ws.cell(row=row_num, column=3, value=result.case_number)
        ws.cell(row=row_num, column=4, value=result.defendant_name)
        ws.cell(row=row_num, column=5, value=result.booking_date)
        ws.cell(row=row_num, column=6, value=result.booking_number)
        ws.cell(row=row_num, column=7, value=result.charges_at_booking)
        ws.cell(row=row_num, column=8, value=result.bond_amount)

        # Look up original charges and case status from defendant
        original_charges = None
        case_status = None
        for d in report.defendants_checked:
            if d.full_name == result.defendant_name and d.case_number == result.case_number:
                original_charges = d.charges
                case_status = d.case_status
                break

        ws.cell(row=row_num, column=9, value=original_charges)
        ws.cell(row=row_num, column=10, value=case_status)

        # Apply highlight coloring
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = highlight_fill
            ws.cell(row=row_num, column=col).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 16  # Mugshot (100px wide images)
    ws.column_dimensions['B'].width = 15  # Matter Number
    ws.column_dimensions['C'].width = 20  # Case Number
    ws.column_dimensions['D'].width = 30  # Defendant Name
    ws.column_dimensions['E'].width = 15  # Booking Date
    ws.column_dimensions['F'].width = 15  # Booking Number
    ws.column_dimensions['G'].width = 40  # Charges
    ws.column_dimensions['H'].width = 15  # Bond Amount
    ws.column_dimensions['I'].width = 40  # Original Charges
    ws.column_dimensions['J'].width = 20  # Case Status

    # Freeze panes at row 4 (after title, empty row, and headers)
    ws.freeze_panes = 'A4'


if __name__ == "__main__":
    import doctest
    doctest.testmod()
