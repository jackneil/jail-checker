"""Validate that mugshots are aligned with data rows."""

from openpyxl import load_workbook

excel_path = r'C:\Github\jail-checker\output\Prosecutor Worklist Report_custody_20251027_165735.xlsx'

print(f"Validating alignment in: {excel_path}")
print()

try:
    wb = load_workbook(excel_path)

    if "In Custody Only" in wb.sheetnames:
        ws = wb["In Custody Only"]
        print("=== IN CUSTODY ONLY SHEET ===")
        print(f"Max row: {ws.max_row}")
        print()

        # Check first 7 rows
        for row_num in range(1, min(8, ws.max_row + 1)):
            print(f"Row {row_num}:")

            # Get first few cell values
            row_values = []
            for col_num in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                val = str(cell.value)[:30] if cell.value else ""
                row_values.append(val)
            print(f"  Data: {row_values}")

            # Check for merged cells
            for merged_range in ws.merged_cells.ranges:
                if row_num >= merged_range.min_row and row_num <= merged_range.max_row:
                    print(f"  [MERGED: {merged_range}]")

            # Check for images
            for img in ws._images:
                if img.anchor._from.row + 1 == row_num:  # openpyxl uses 0-indexed internally
                    print(f"  [IMAGE ANCHORED HERE]")

            print()

        print(f"\nTotal images in sheet: {len(ws._images)}")
        print("\nImage anchor summary:")
        for i, img in enumerate(ws._images[:5], 1):
            row = img.anchor._from.row + 1  # Convert to 1-indexed
            col = img.anchor._from.col
            print(f"  Image {i}: anchored at row {row}, col {col}")

    wb.close()
    print("\n✅ Validation complete")

except Exception as e:
    print(f"ERROR: {type(e).__name__}: {str(e)}")
    import traceback
    traceback.print_exc()
