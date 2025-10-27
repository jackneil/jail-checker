"""Check the layout of an old Excel file to understand the original structure."""

from openpyxl import load_workbook

# Check the first generated file
old_file = r'C:\Github\jail-checker\output\Prosecutor Worklist Report_custody_20251027_160653.xlsx'

print(f"Checking old file layout: {old_file}")
print()

try:
    wb = load_workbook(old_file)

    if "In Custody Only" in wb.sheetnames:
        ws = wb["In Custody Only"]
        print("=== IN CUSTODY ONLY SHEET ===")
        print(f"Max row: {ws.max_row}")
        print()

        # Check first 5 rows
        for row_num in range(1, min(6, ws.max_row + 1)):
            print(f"Row {row_num}:")
            row_values = []
            for col_num in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                row_values.append(str(cell.value)[:20] if cell.value else "")
            print(f"  {row_values}")

            # Check if row is merged
            for merged_range in ws.merged_cells.ranges:
                if row_num >= merged_range.min_row and row_num <= merged_range.max_row:
                    print(f"  [MERGED: {merged_range}]")
            print()

    wb.close()

except Exception as e:
    print(f"ERROR: {type(e).__name__}: {str(e)}")
    import traceback
    traceback.print_exc()
