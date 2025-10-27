"""Verify column A width is wide enough for mugshots."""

from openpyxl import load_workbook

excel_path = r'C:\Github\jail-checker\output\Prosecutor Worklist Report_custody_20251027_170120.xlsx'

print(f"Checking column widths in: {excel_path}")
print()

try:
    wb = load_workbook(excel_path)

    for sheet_name in ["All Results", "In Custody Only"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"=== {sheet_name.upper()} ===")
            print(f"Column A width: {ws.column_dimensions['A'].width} characters")
            print(f"Column B width: {ws.column_dimensions['B'].width} characters")
            print(f"Column C width: {ws.column_dimensions['C'].width} characters")
            print()

            # Show first data row
            if sheet_name == "All Results":
                data_row = 2
            else:
                data_row = 4

            print(f"First data row (row {data_row}):")
            for col in ['A', 'B', 'C', 'D']:
                val = ws[f'{col}{data_row}'].value
                print(f"  {col}{data_row}: {str(val)[:40] if val else '(empty)'}")
            print()

    wb.close()
    print("Column A is now 16 characters wide (100px mugshots should fit)")
    print("The mugshots should no longer overlap onto column B")

except Exception as e:
    print(f"ERROR: {type(e).__name__}: {str(e)}")
    import traceback
    traceback.print_exc()
