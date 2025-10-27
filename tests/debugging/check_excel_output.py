"""Quick script to check if Excel output has all the new fields populated."""

import openpyxl

wb = openpyxl.load_workbook(r'C:\Github\jail-checker\output\Prosecutor Worklist Report_custody_20251027_160653.xlsx')
ws = wb['All Results']

print('=== Headers ===')
headers = [cell.value for cell in ws[1]]
print(headers)
print()

print('=== First IN CUSTODY Record (Looking for Murray) ===')
for row in ws.iter_rows(min_row=2, max_row=25, values_only=True):
    if row[3] and 'IN CUSTODY' in str(row[3]) and 'Murray' in str(row[2]):
        print(f'Matter: {row[0]}')
        print(f'Case: {row[1]}')
        print(f'Name: {row[2]}')
        print(f'Status: {row[3]}')
        print(f'Booking Date: {row[4]}')
        print(f'Booking Number: {row[5]}')
        print(f'Charges: {row[6]}')
        print(f'Bond: {row[7]}')
        break

print()
print('=== All IN CUSTODY records with their booking dates and bonds ===')
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] and 'IN CUSTODY' in str(row[3]):
        name = row[2]
        booking_date = row[4]
        bond = row[7]
        print(f'{name:<30} | Booking: {booking_date or "MISSING":<12} | Bond: {bond or "MISSING"}')
