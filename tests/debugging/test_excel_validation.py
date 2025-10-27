"""Test to verify the generated Excel file is not corrupt."""

from openpyxl import load_workbook
import sys

excel_path = r'C:\Github\jail-checker\output\Prosecutor Worklist Report_custody_20251027_165236.xlsx'

print(f"Testing Excel file: {excel_path}")
print()

try:
    print("Loading workbook...")
    wb = load_workbook(excel_path)
    print(f"SUCCESS: Workbook loaded")
    print(f"Sheet names: {wb.sheetnames}")
    print()

    # Check All Results sheet
    if "All Results" in wb.sheetnames:
        ws = wb["All Results"]
        print(f"All Results sheet:")
        print(f"  Max row: {ws.max_row}")
        print(f"  Max column: {ws.max_column}")
        print(f"  Headers: {[cell.value for cell in ws[1]]}")
        print()

        # Count images
        image_count = len(ws._images)
        print(f"  Images in sheet: {image_count}")
        print()

    # Check In Custody sheet
    if "In Custody" in wb.sheetnames:
        ws = wb["In Custody"]
        print(f"In Custody sheet:")
        print(f"  Max row: {ws.max_row}")
        print(f"  Max column: {ws.max_column}")
        print(f"  Headers: {[cell.value for cell in ws[2]]}")  # Row 2 has headers
        print()

        # Count images
        image_count = len(ws._images)
        print(f"  Images in sheet: {image_count}")
        print()

    wb.close()
    print("File validated successfully - NO CORRUPTION")
    sys.exit(0)

except Exception as e:
    print(f"ERROR: {type(e).__name__}: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
