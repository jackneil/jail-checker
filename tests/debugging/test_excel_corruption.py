"""Test to identify Excel corruption issue with mugshots."""

import sys
sys.path.insert(0, r'C:\Github\jail-checker\src')

from io import BytesIO
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

# Test with actual mugshot URL
test_url = "https://analytics.southernsoftware.com/crons/CCMugshotsNameID/SC018013C-213349.jpg"

print("Creating test workbook with mugshot...")

wb = Workbook()
ws = wb.active
ws.title = "Test"

# Download image
response = requests.get(test_url, timeout=10)
image_data = BytesIO(response.content)

# IMPORTANT: Seek to beginning before creating Image
image_data.seek(0)

# Create Image
img = Image(image_data)
img.width = 100
img.height = 125
img.anchor = 'A1'

# Set row height
ws.row_dimensions[1].height = 100

# Add some data
ws['B1'] = "Name"
ws['C1'] = "Status"

# Add image
ws.add_image(img)

# Save
output_path = r'C:\Github\jail-checker\tests\debugging\test_corruption.xlsx'
print(f"Saving to: {output_path}")
wb.save(output_path)
print("File saved successfully")

# Try to open it to verify it's not corrupt
print("\nAttempting to reload the file...")
try:
    wb_reload = load_workbook(output_path)
    print(f"SUCCESS: File loaded successfully")
    print(f"Sheet names: {wb_reload.sheetnames}")
    ws_reload = wb_reload.active
    print(f"Cell B1 value: {ws_reload['B1'].value}")
    wb_reload.close()
except Exception as e:
    print(f"ERROR loading file: {type(e).__name__}: {str(e)}")
    import traceback
    traceback.print_exc()
